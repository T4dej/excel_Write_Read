import openpyxl

original = "D:\\directory\\original.xlsx"  #path to original file that we are updating

wb_original = openpyxl.load_workbook(original)
original = wb_original.active
row_original = original.max_row

new_file = "D:\\directory\\newFile.xlsx"   #path to new file, from where we are geting data

wb_file = openpyxl.load_workbook(new_file)
file = wb_file.active
row_file = file.max_row

data = []

for z in range(1, row_original+1, 1):    #we put all data from new file to a list
    metaData = original.cell(row=z, column=2).value
    data.append(metaData)


for i in range(1, row_file+1, 1):
    number_id = file.cell(row=i, column=2).value   #here we choose the column with values that we are interested in
    points = file.cell(row=i, column=3).value      #here we choose the column with values that we are interested in
    if number_id in data:    #if person is in the list we proceed with the update
        for j in range(1, row_original, 1):  # looping over second file
            number_id_search = original.cell(row=j, column=2).value
            if number_id_search == number_id:  # searching for persone to be updated
                original.cell(row=j, column=3).value += points  # adding points to that person
                break
    if number_id not in data:      #if there is no person in the original table, we are adding the whole row to original file
        row_original += 1
        for row in range(i, row_file+1, 1):
            for column in range(1, file.max_column + 1, 1):
                original.cell(row=row_original, column=column).value = file.cell(row=i, column=column).value

wb_original.save("D:\\directory\\newOriginal.xlsx")   #path to new saved original file